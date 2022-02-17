from __future__ import unicode_literals
from os import stat
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _, log_error
from frappe.utils.csvutils import UnicodeWriter, build_csv_response, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'Form-27'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    header = title1(args)   
    ws.append(header)
    header = title2(args)   
    ws.append(header)
    header = title3(args)   
    ws.append(header)
    header = title4(args)   
    ws.append(header)
    header = title6(args)
    ws.append(header)
    header = title5(args)   
    ws.append(header)
    
     
    data = get_data(args)

    for row in data:
        ws.append(row)

    ws.sheet_view.zoomScale = 80 

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def title1(args):
    contractor_name = '-'
    if args.contractor :
        contractor_name = args.contractor
    data=['Name and address of the contractor: ',]
    data.append(contractor_name)
    return data

@frappe.whitelist()
def title2(args):
    data=['Name and addres of establishment in /under which contract is carried on:',]
    return data

@frappe.whitelist()
def title3(args):
    data=['Nature and location of work ',]
    data.append('Tamilnadu Petroproducts Limited')
    return data

@frappe.whitelist()
def title4(args):
    data=['Wage Period Monthly',]
    return data

@frappe.whitelist()
def title6(args):
    data=['Name and Address of the Principal Employer',]
    return data

@frappe.whitelist()
def title5(args):
    data = ['Sr no','Name of workman','Serial No in the register of workman','Designation/nature of workdone','No of days worked','Units of workdone','Daily rate of wages/piece rate','Basic wages','Dearness Allowance','Overtime','Other cash payments','Total','Deduction if any','Net amount paid','Signature/Thumb impression o workman','Initial of contractor the representative']
    return data


def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates

@frappe.whitelist()
def get_data(args):
    data=[]
    values = frappe.db.sql("select name,`tabSalary Slip`.employee_name as employee_name,`tabSalary Slip`.employee as employee,`tabSalary Slip`.payment_days as payment_days,`tabSalary Slip`.rounded_total as rounded,`tabSalary Slip`.total_deduction as deduction,`tabSalary Slip`.net_pay as net_pay from `tabSalary Slip` where `tabSalary Slip`.start_date between '%s'and '%s'"%(args.from_date,args.to_date),as_dict=True)
    if args.contractor:
        values = frappe.db.sql("select name,`tabSalary Slip`.employee_name as employee_name,`tabSalary Slip`.employee as employee,`tabSalary Slip`.payment_days as payment_days,`tabSalary Slip`.rounded_total as rounded,`tabSalary Slip`.total_deduction as deduction,`tabSalary Slip`.net_pay as net_pay from `tabSalary Slip` where `tabSalary Slip`.contractor = '%s' and `tabSalary Slip`.start_date between '%s'and '%s'"%(args.contractor,args.from_date,args.to_date),as_dict=True)

    i = 1
    for value in values:
        row = []
        row.extend([i,value.employee_name,value.employee,'',value.payment_days,'',540])
        basic = frappe.db.get_value("Salary Detail",{'abbr':'B','parent':value.name},'amount')
        dearness = frappe.db.get_value("Salary Detail",{'abbr':'DA','parent':value.name},'amount')
        overtime = frappe.db.get_value("Salary Detail",{'abbr':'OT','parent':value.name},'amount')

        row.append(basic)
        row.append(dearness)
        row.append(overtime)
        row.append('')
        row.append(value.rounded)
        row.append(value.deduction)
        row.append(value.net_pay)
        data.append(row)
        i +=1
    return data
