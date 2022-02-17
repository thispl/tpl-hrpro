from __future__ import unicode_literals
from os import error, stat
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _, log_error
from frappe.utils.csvutils import UnicodeWriter, build_csv_response, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'Form-26'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    header = add_title1(args)   
    ws.append(header)
    header2=add_title2(args)
    ws.append(header2)
    header3= add_title3(args)
    ws.append(header3)
    ws.append([''])
    header_day = add_day(args)
    ws.append(header_day)
   
   
   
    data = get_data(args)

    for row in data:
        ws.append(row)

    ws.sheet_view.zoomScale = 80 
    employee = frappe.get_all('Employee',{'employment_type':args.employment_type,'status':'Active'},['name','employee_name','department','category'])
    if args.contractor:
        employee = frappe.get_all('Employee',{'employment_type':args.employment_type,'status':'Active','contractor_id':args.contractor},['name','employee_name','department','category'])
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=len(employee)+6, min_col=1, max_col=45):
        for cell in rows:
            cell.border = border
    for rows in ws.iter_rows(min_row=1, max_row=len(employee)+5, min_col=1, max_col=35):
        for cell in rows:
            # if cell.value:
            #     frappe.log_error(message=cell.value)
            if cell.value == 'A':
                cell.fill = PatternFill(fgColor='FF0000', fill_type = "solid")

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def add_day(args):
    data = ['Employee ID',"Employee name",'Department','Category']
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%d %a')
        data.append(day_format)
    data.append('No of Days Worked')
    data.append('No of Days Worked (as per Contractor)')
    data.append('Working Days Diff')
    data.append('Total Wages')
    data.append('PF')
    data.append('ESI')
    data.append('Net Salary')
    data.append('PF(as per Contractor)')
    data.append('ESI(as per Contractor)')
    data.append('PF Diff')
    data.append('ESIC Diff')
    data.append('Total Diff')
    return data    


@frappe.whitelist()
def add_title1(args):
    data = ['Name and address of the Employer :','Tamilnadu Petroproducts Limited' ,'','','','','Month :']
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%b')
    data.append(day_format)
    
    return data


@frappe.whitelist()
def add_title2(args):
    contractor_name = '-'
    if args.contractor :
        contractor_name = args.contractor
        
    data = ['Name and address of the Contractor : ',contractor_name, '','','','','Year : ']
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%Y')
    data.append(day_format)
    return data

@frappe.whitelist()
def add_title3(args):
    data = ['Name and Location of the Working Site : ']
    data.append(args.plant)
    return data

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates

def get_data(args):
    data=[]
    row=[]
    dates=get_dates(args)
    employee = frappe.get_all('Employee',{'employment_type':args.employment_type,'status':'Active'},['name','employee_name','department','category'])
    if args.contractor:
        employee = frappe.get_all('Employee',{'employment_type':args.employment_type,'status':'Active','contractor_id':args.contractor},['name','employee_name','department','category'])
    for emp in employee:
        row = ([emp.name,emp.employee_name,emp.department or '-',emp.category or '-'])
        no_of_days = 0
        for date in dates:
            if args.plant == 'TPL (All Plants)':
                status=frappe.db.get_value('Attendance',{'employee':emp.name,'attendance_date':date,},'status')
            else:
                status=frappe.db.get_value('Attendance',{'employee':emp.name,'attendance_date':date,'plant':args.plant,},'status')
            if status == 'Present':
                status = 'P'
            elif status == 'Half Day':
                status = 'P'
            else:
                status ='A'
            row.append(status)
            if args.plant == 'TPL (All Plants)':
                work_status = frappe.db.get_value('Attendance',{'employee':emp.name,'attendance_date':date},'status')
            else:
                work_status = frappe.db.get_value('Attendance',{'employee':emp.name,'attendance_date':date,'plant':args.plant,},'status')
            if work_status == 'Present':
                no_of_days += 1
            elif work_status == 'Half Day':
                no_of_days += 1
            else:
                no_of_days += 0
        total_days_worked = no_of_days
        basic = frappe.db.get_value('Employee',{'employee':emp.name},'basic')
        DA = frappe.db.get_value('Employee',{'employee':emp.name},'da')
        cont_pf = frappe.db.get_value('Contractor PF ESI',{'employee':emp.name,'payroll_date':args.from_date},'pf') or 0
        cont_esi = frappe.db.get_value('Contractor PF ESI',{'employee':emp.name,'payroll_date':args.from_date},'esi') or 0
        cont_working_days = frappe.db.get_value('Contractor PF ESI',{'employee':emp.name,'payroll_date':args.from_date},'working_days') or 0
        days_diff = total_days_worked - cont_working_days
        basic_da = basic+DA
        total_wages = basic_da * total_days_worked
        pf = total_wages*0.12
        esi = total_wages * 0.0075
        net_salary = total_wages - pf - esi
        pf_diff = pf - cont_pf
        esi_diff = esi - cont_esi 
        total_diff = pf_diff + esi_diff
        row.append(total_days_worked)
        row.append(cont_working_days)
        row.append(days_diff)
        row.append(total_wages)
        row.append(pf)
        row.append(esi)
        row.append(net_salary)
        row.append(cont_pf)
        row.append(cont_esi)
        row.append(pf_diff)
        row.append(esi_diff)
        row.append(total_diff)
        data.append(row)   
    tot = total(args)
    data.append(tot)
    return data

def total(args):
    data =['TOTAL','','','']
    dates =get_dates(args)
    employee = frappe.get_all('Employee',{'employment_type':args.employment_type,'status':'Active'},['name','employee_name',])

    if args.contractor:
        employee = frappe.get_all('Employee',{'employment_type':args.employment_type,'status':'Active','contractor_id':args.contractor},['name','employee_name',])
    
    emp_list =[]
    for emp in employee:
        emp_list.append(emp.name)
    for date in dates:
        if args.plant == 'TPL (All Plants)':
            status_count = frappe.db.count('Attendance',{'employee':('in',(emp_list)),'attendance_date':date,'status':('in',['Present','Half Day'])})
        else:
            status_count = frappe.db.count('Attendance',{'employee':('in',(emp_list)),'attendance_date':date,'plant':args.plant,'status':('in',['Present','Half Day'])})
      
        data.append(status_count)


    # if args.plant == 'TPL (All Plants)':
    #     total_work = frappe.db.count('Attendance',{'employee':('in',(emp_list)),'attendance_date':('between',(args.from_date and args.to_date)),'status':('in',['Present','Half Day'])})
    # else:
    #     total_work = frappe.db.count('Attendance',{'employee':('in',(emp_list)),'attendance_date':('between',(args.from_date and args.to_date)),'plant':args.plant,'status':('in',['Present','Half Day'])})
    #     frappe.error_log()
    # if args.plant == 'TPL (All Plants)':
    #     total_work = frappe.db.sql( f'select count(*) as count from `tabAttendance` where employee in {emp_list}, attendance_date between {args.from_date} and {args.to_date}, status in (Present,Half Day)')
    # else:
    #     total_work = frappe.db.sql( f'select count(*) as count from `tabAttendance` where employee in {emp_list}, attendance_date between {args.from_date} and {args.to_date},plant =  {args.plant},status in (Present,Half Day)')

    data.append("Total No of Days")
   

   
    return data
    
    